# -*- coding: utf-8 -*-
"""
Running this 
"""

from openpyxl import load_workbook
import pandas as pd

def get_coeffs(wb):
    '''Returns a dictionary with data from named ranges specified by names_list
    from the xls passed as wb'''

    # First get the names from the target sheet. 
    # wb.defined_names is an object that has this info.
    # Probably a better way than this to extract names but hey
    names_list = [x.split(",")[0][1:].split("'")[0] 
                    for x in str(wb.defined_names).split('name=')
                    if "xlnm" not in x][1:]

    # Now load up the data - all named ranges    
    raw_dict = {}
    for name in names_list:
        raw_dict[name] = list(wb.defined_names[name].destinations)[0]

    # Now put them in a dictionary
    coeffs = {}
    for sheet, rng in raw_dict: 
        ws = wb[sheet]
        vals = []

        # iterate through the range if it's more than one value, otherwise just take the single value
        if type(ws[rng]) is tuple:
            for cell in ws[rng]:
                vals.append(cell[0].value)
        else:
            vals.append(ws[rng].value)
        
        coeffs[coeff] = vals
    
    return coeffs

def make_pdseries(in_dict):
    '''Turns a dictionary into a pandas Series'''
    ser = pd.Series()
    # populate by appending to the empty series.  Note index of tuples
    for key in in_dict:
        for i,datum in enumerate(in_dict[key]):
            ser = ser.append(pd.Series({(key,i):datum}))
    # now just turn the index into a proper multi-index
    ser.index = pd.MultiIndex.from_tuples(ser.index)
    return ser


if __name__=='__main__':
    # USAGE
    # First load in the workbook (data only, i.e. not formulas)
    import sys
    sys.argv[1]='target'
    wb = load_workbook('target', data_only=True)
    
    # Finally run it
    coeff_ser = make_pdseries(get_coeffs(wb))
    